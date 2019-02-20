# -*- coding: utf-8 -*-
"""
Created on Thu Feb 14 09:32:00 2019

@author: ISP_60769
"""

import openpyxl
import webbrowser
import requests,bs4
import csv

# ワークブック読み込み
wb = openpyxl.load_workbook('excel_web.xlsx')

# シートを指定
#sheet = wb.get_sheet_by_name('Sheet1')
sheet = wb.active

# シートからセルを取得
#list(sheet.columns)[0]
print(sheet.max_row)
print(sheet.max_column)
for cell_obj in list(sheet.columns)[0]:
    print(cell_obj.value)
    
    #webページを開く
    #webbrowser.open(cell_obj.value)
    
    #webページの内容を取得
    #res = requests.get('http://inventwithpython.com/')
    res = requests.get(cell_obj.value, verify=False)
    try:
        res.raise_for_status()
        print('OK')
        
        # web 内容抜き出し
        soup = bs4.BeautifulSoup(res.text, "html.parser")
        
        # file open
#        play_file=open('test1.txt','wb')
        csv_file=open('output.csv','w',newline='')
        output_writer=csv.writer(csv_file)
        
        # class get
        card_t = soup.select('.card-title')
        
        for card in card_t:
 #           out_id = card.get('id')
            out_card=card.getText()
            output_writer.writerow([0,out_card])
            print(out_card)
            
        
        csv_file.close()
        
        #web内容をファイルに書き出し
#        for chunk in res.iter_content(100000):
#            play_file.write(chunk)
#        play_file.close()
    #requestsの例外処理
    except Exception as exc:
        print('問題あり:{}'.format(exc))